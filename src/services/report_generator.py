"""Service to generate a Profit & Loss (P&L) financial report."""

import logging
from decimal import Decimal
from pathlib import Path
from typing import BinaryIO, Dict, List, Set

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from src.models.asset import Asset
from src.models.report import Report, ReportLineItem, DisallowedExpenseItem
from src.models.transaction import Transaction
from src.services.depreciation import get_depreciation_for_year

logger = logging.getLogger(__name__)

# Categories that are excluded from P&L (owner withdrawals, corrections, etc.)
EXCLUDED_CATEGORY_TYPES = {'excluded'}


class ReportGenerator:
    """Generates a P&L report from transactions and assets."""

    def __init__(self, transactions: List[Transaction], assets: List[Asset],
                 categories: Dict = None):
        self.transactions = transactions
        self.assets = assets
        self.categories = categories or {}
        self._excluded_categories: Set[str] = self._get_excluded_categories()

    def _get_excluded_categories(self) -> Set[str]:
        """Get set of category IDs that should be excluded from P&L."""
        excluded = set()
        for cat_id, cat in self.categories.items():
            if hasattr(cat, 'type') and cat.type in EXCLUDED_CATEGORY_TYPES:
                excluded.add(cat_id)
        # Always exclude these known categories
        excluded.update({'prive-opname', 'loon', 'verkeerde-rekening',
                        'interne-storting', 'mastercard'})
        return excluded

    def _get_partially_deductible_categories(self) -> Dict[str, int]:
        """Get categories with partial deductibility and their percentages."""
        partial = {}
        for cat_id, cat in self.categories.items():
            if hasattr(cat, 'deductibility_pct') and cat.deductibility_pct < 100:
                partial[cat_id] = cat.deductibility_pct
        return partial

    def generate_pnl_report(self, fiscal_year: int) -> Report:
        """
        Generates a P&L report for a given fiscal year.
        """
        report = Report(fiscal_year=fiscal_year)
        logger.debug(f"Generating P&L report for fiscal year {fiscal_year} with {len(self.transactions)} transactions and {len(self.assets)} assets.")
        
        # Filter transactions for the fiscal year
        year_transactions = [
            t for t in self.transactions if t.booking_date.year == fiscal_year
        ]
        logger.debug(f"Found {len(year_transactions)} transactions for {fiscal_year}.")

        if not year_transactions:
            return report

        df = pd.DataFrame([t.to_dict() for t in year_transactions])
        df['amount'] = df['amount'].apply(Decimal)
        logger.debug("Transaction DataFrame head:\n%s", df.head().to_string())

        # Handle income
        income_df = df[df['category'] == 'omzet']
        if not income_df.empty:
            total_omzet = income_df['amount'].sum()
            omzet_item = ReportLineItem(category='omzet', amount=total_omzet, item_type='income')
            logger.debug(f"Total income (omzet): {total_omzet}")

            thera_amount = income_df[income_df['is_therapeutic'] == True]['amount'].sum()
            non_thera_amount = income_df[income_df['is_therapeutic'] == False]['amount'].sum()
            logger.debug(f"Therapeutic: {thera_amount}, Non-therapeutic: {non_thera_amount}")

            if thera_amount > 0:
                omzet_item.sub_items.append(ReportLineItem('Therapeutisch (BTW-vrijgesteld art. 44)', thera_amount, 'income'))
            if non_thera_amount > 0:
                omzet_item.sub_items.append(ReportLineItem('Niet-therapeutisch', non_thera_amount, 'income'))
            
            report.income_items.append(omzet_item)

        # Handle expenses (excluding owner withdrawals and corrections)
        expense_df = df[
            df['category'].notna() &
            (df['category'] != 'omzet') &
            (~df['category'].isin(self._excluded_categories))
        ]
        if not expense_df.empty:
            expense_groups = expense_df.groupby('category')['amount'].sum()
            logger.debug("Expense groups:\n%s", expense_groups.to_string())
            for category, total in expense_groups.items():
                report.expense_items.append(ReportLineItem(category, total, 'expense'))

        # Handle depreciation
        depreciation_entries = get_depreciation_for_year(self.assets, fiscal_year)
        if depreciation_entries:
            total_depreciation = sum(e.amount for e in depreciation_entries)
            logger.debug(f"Total depreciation: {total_depreciation}")
            # Depreciation is an expense, so it should be negative
            report.depreciation_items.append(ReportLineItem('afschrijvingen', -total_depreciation, 'depreciation'))

        # Handle uncategorized
        uncategorized_df = df[df['category'].isna()]
        if not uncategorized_df.empty:
            total_uncategorized = uncategorized_df['amount'].sum()
            logger.debug(f"Total uncategorized: {total_uncategorized}")
            report.uncategorized_items.append(ReportLineItem('Uncategorized', total_uncategorized, 'uncategorized'))

        # Handle verkeerde-rekening (private expense) transactions for warning tracking
        verkeerde_df = df[df['category'] == 'verkeerde-rekening']
        if not verkeerde_df.empty:
            # Create individual line items for each transaction (needed for detailed listing)
            for _, row in verkeerde_df.iterrows():
                report.verkeerde_rekening_items.append(
                    ReportLineItem(
                        category='verkeerde-rekening',
                        amount=row['amount'],
                        item_type='expense'
                    )
                )
            logger.debug(f"Verkeerde-rekening transactions: {len(report.verkeerde_rekening_items)}, "
                        f"balance: {report.verkeerde_rekening_balance}")

        # Calculate disallowed expenses (verworpen uitgaven)
        partial_categories = self._get_partially_deductible_categories()
        for category_id, deductible_pct in partial_categories.items():
            cat_df = df[df['category'] == category_id]
            if not cat_df.empty:
                total_amount = abs(cat_df['amount'].sum())  # Expenses are negative
                deductible_amount = total_amount * Decimal(deductible_pct) / 100
                disallowed_amount = total_amount - deductible_amount

                cat_name = category_id
                if category_id in self.categories:
                    cat_name = self.categories[category_id].name

                report.disallowed_expenses.append(DisallowedExpenseItem(
                    category=cat_name,
                    total_amount=total_amount,
                    deductible_pct=deductible_pct,
                    deductible_amount=deductible_amount,
                    disallowed_amount=disallowed_amount
                ))
                logger.debug(f"Verworpen uitgave: {cat_name} - Total: {total_amount}, "
                           f"Deductible: {deductible_pct}%, Disallowed: {disallowed_amount}")

        return report

    def export_to_excel(self, report: Report, output_path: Path | BinaryIO) -> None:
        """
        Exports the P&L report to a formatted Excel file with transactions tab.
        """
        from openpyxl.styles import PatternFill, Border, Side

        logger.info(f"Starting Excel export to {output_path}")
        wb = Workbook()
        ws = wb.active
        ws.title = f"P&L {report.fiscal_year}"

        # Define styles
        bold_font = Font(bold=True)
        currency_format = '€ #,##0.00'
        header_fill = PatternFill(start_color="2c5282", end_color="2c5282", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.cell(row=1, column=1, value=f"Resultatenrekening {report.fiscal_year}").font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')

        current_row = 3
        logger.debug("Writing income section...")
        # Income (Baten)
        ws.cell(row=current_row, column=1, value="Baten").font = bold_font
        current_row += 1
        for item in report.income_items:
            ws.cell(row=current_row, column=1, value=f"  {item.category}")
            cell = ws.cell(row=current_row, column=2, value=float(item.amount))
            cell.number_format = currency_format
            current_row += 1
            for sub_item in item.sub_items:
                ws.cell(row=current_row, column=1, value=f"    - {sub_item.category}")
                cell = ws.cell(row=current_row, column=2, value=float(sub_item.amount))
                cell.number_format = currency_format
                current_row += 1

        ws.cell(row=current_row, column=1, value="Totaal Baten").font = bold_font
        cell = ws.cell(row=current_row, column=2, value=float(report.total_income))
        cell.number_format = currency_format
        cell.font = bold_font
        current_row += 2

        # Kosten (operational expenses - NOT depreciation)
        logger.debug("Writing kosten section...")
        ws.cell(row=current_row, column=1, value="Kosten").font = bold_font
        current_row += 1
        total_kosten = Decimal(0)
        for item in sorted(report.expense_items, key=lambda x: x.category):
            ws.cell(row=current_row, column=1, value=f"  {item.category.replace('-', ' ').title()}")
            cell = ws.cell(row=current_row, column=2, value=float(-item.amount))
            cell.number_format = currency_format
            total_kosten += -item.amount
            current_row += 1

        ws.cell(row=current_row, column=1, value="Totaal Kosten").font = bold_font
        cell = ws.cell(row=current_row, column=2, value=float(total_kosten))
        cell.number_format = currency_format
        cell.font = bold_font
        current_row += 2

        # Afschrijvingen (Depreciation - separate from Kosten)
        if report.depreciation_items:
            logger.debug("Writing afschrijvingen section...")
            ws.cell(row=current_row, column=1, value="Afschrijvingen").font = bold_font
            current_row += 1

            # Show asset details
            total_afschrijvingen = Decimal(0)
            from src.services.depreciation import get_depreciation_for_year
            depreciation_entries = get_depreciation_for_year(self.assets, report.fiscal_year)
            for entry in depreciation_entries:
                asset_name = entry.asset_id
                for asset in self.assets:
                    if asset.id == entry.asset_id:
                        asset_name = asset.name
                        break
                ws.cell(row=current_row, column=1, value=f"  {asset_name}")
                cell = ws.cell(row=current_row, column=2, value=float(entry.amount))
                cell.number_format = currency_format
                total_afschrijvingen += entry.amount
                current_row += 1

            ws.cell(row=current_row, column=1, value="Totaal Afschrijvingen").font = bold_font
            cell = ws.cell(row=current_row, column=2, value=float(total_afschrijvingen))
            cell.number_format = currency_format
            cell.font = bold_font
            current_row += 2

        # Result
        logger.debug("Writing result section...")
        ws.cell(row=current_row, column=1, value="Resultaat (Winst/Verlies)").font = bold_font
        cell = ws.cell(row=current_row, column=2, value=float(report.profit_loss))
        cell.number_format = currency_format
        cell.font = bold_font
        current_row += 2

        # Verworpen uitgaven (Disallowed Expenses)
        if report.disallowed_expenses:
            logger.debug("Writing verworpen uitgaven section...")
            ws.cell(row=current_row, column=1, value="Verworpen Uitgaven (Disallowed Expenses)").font = bold_font
            current_row += 1

            # Header row for the table
            ws.cell(row=current_row, column=1, value="Categorie")
            ws.cell(row=current_row, column=2, value="Totaal bedrag")
            ws.cell(row=current_row, column=3, value="% Aftrekbaar")
            ws.cell(row=current_row, column=4, value="Aftrekbaar")
            ws.cell(row=current_row, column=5, value="Verworpen")
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).font = Font(italic=True)
            current_row += 1

            for item in sorted(report.disallowed_expenses, key=lambda x: x.category):
                ws.cell(row=current_row, column=1, value=f"  {item.category}")
                cell = ws.cell(row=current_row, column=2, value=float(item.total_amount))
                cell.number_format = currency_format
                ws.cell(row=current_row, column=3, value=f"{item.deductible_pct}%")
                cell = ws.cell(row=current_row, column=4, value=float(item.deductible_amount))
                cell.number_format = currency_format
                cell = ws.cell(row=current_row, column=5, value=float(item.disallowed_amount))
                cell.number_format = currency_format
                current_row += 1

            ws.cell(row=current_row, column=1, value="Totaal Verworpen Uitgaven").font = bold_font
            cell = ws.cell(row=current_row, column=5, value=float(report.total_disallowed))
            cell.number_format = currency_format
            cell.font = bold_font
            current_row += 2

        # Uncategorized warning
        if report.total_uncategorized != 0:
            logger.debug("Adding uncategorized warning.")
            ws.cell(row=current_row, column=1, value=f"Waarschuwing: {report.total_uncategorized:,.2f} aan niet-gecategoriseerde transacties (niet opgenomen in resultaat).").font = Font(color="FF0000")
            current_row += 1

        # Verkeerde-rekening warning (US2)
        if report.verkeerde_rekening_balance != 0:
            logger.debug("Adding verkeerde-rekening warning.")
            balance = float(report.verkeerde_rekening_balance)
            sign = "+" if balance > 0 else ""
            ws.cell(row=current_row, column=1, value=f"Waarschuwing: Verkeerde rekening niet in balans (€ {sign}{balance:,.2f}).").font = Font(color="FF0000")
            current_row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # Create Transactions sheet
        logger.debug("Creating transactions sheet...")
        ws_tx = wb.create_sheet(title="Verrichtingen")

        # Filter transactions for this fiscal year
        year_transactions = [
            t for t in self.transactions
            if t.booking_date.year == report.fiscal_year and not t.is_excluded
        ]
        # Sort by date
        year_transactions.sort(key=lambda t: t.booking_date)

        # Headers - retain all columns from input
        headers = [
            'Rekening', 'Boekingsdatum', 'Valutadatum', 'Rekeninguittrekselnr', 'Transactienr',
            'Tegenpartij', 'Rekening Tegenpartij', 'Straat en nummer', 'Postcode en plaats',
            'BIC', 'Landcode', 'Omschrijving', 'Bedrag', 'Devies',
            'Categorie', 'Therapeutisch', 'Bron'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_tx.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_num, tx in enumerate(year_transactions, 2):
            ws_tx.cell(row=row_num, column=1, value=tx.own_account or '').border = thin_border
            ws_tx.cell(row=row_num, column=2, value=tx.booking_date.strftime('%d/%m/%Y')).border = thin_border
            ws_tx.cell(row=row_num, column=3, value=tx.value_date.strftime('%d/%m/%Y')).border = thin_border
            ws_tx.cell(row=row_num, column=4, value=tx.statement_number or '').border = thin_border
            ws_tx.cell(row=row_num, column=5, value=tx.transaction_number or '').border = thin_border
            ws_tx.cell(row=row_num, column=6, value=tx.counterparty_name or '').border = thin_border
            ws_tx.cell(row=row_num, column=7, value=tx.counterparty_iban or '').border = thin_border
            ws_tx.cell(row=row_num, column=8, value=tx.counterparty_street or '').border = thin_border
            ws_tx.cell(row=row_num, column=9, value=tx.counterparty_postal_city or '').border = thin_border
            ws_tx.cell(row=row_num, column=10, value=tx.counterparty_bic or '').border = thin_border
            ws_tx.cell(row=row_num, column=11, value=tx.counterparty_country or '').border = thin_border
            ws_tx.cell(row=row_num, column=12, value=tx.description or '').border = thin_border

            amount_cell = ws_tx.cell(row=row_num, column=13, value=float(tx.amount))
            amount_cell.number_format = currency_format
            amount_cell.border = thin_border

            ws_tx.cell(row=row_num, column=14, value=tx.currency or 'EUR').border = thin_border
            ws_tx.cell(row=row_num, column=15, value=tx.category or '').border = thin_border
            ws_tx.cell(row=row_num, column=16, value='Ja' if tx.is_therapeutic else 'Nee').border = thin_border
            ws_tx.cell(row=row_num, column=17, value=tx.source_file or '').border = thin_border

        # Adjust column widths
        ws_tx.column_dimensions['A'].width = 22  # Rekening
        ws_tx.column_dimensions['B'].width = 12  # Boekingsdatum
        ws_tx.column_dimensions['C'].width = 12  # Valutadatum
        ws_tx.column_dimensions['D'].width = 18  # Rekeninguittrekselnr
        ws_tx.column_dimensions['E'].width = 12  # Transactienr
        ws_tx.column_dimensions['F'].width = 35  # Tegenpartij
        ws_tx.column_dimensions['G'].width = 24  # Rekening Tegenpartij
        ws_tx.column_dimensions['H'].width = 25  # Straat en nummer
        ws_tx.column_dimensions['I'].width = 22  # Postcode en plaats
        ws_tx.column_dimensions['J'].width = 12  # BIC
        ws_tx.column_dimensions['K'].width = 10  # Landcode
        ws_tx.column_dimensions['L'].width = 50  # Omschrijving
        ws_tx.column_dimensions['M'].width = 14  # Bedrag
        ws_tx.column_dimensions['N'].width = 8   # Devies
        ws_tx.column_dimensions['O'].width = 25  # Categorie
        ws_tx.column_dimensions['P'].width = 12  # Therapeutisch
        ws_tx.column_dimensions['Q'].width = 25  # Bron

        # Add autofilter
        ws_tx.auto_filter.ref = f"A1:Q{len(year_transactions) + 1}"

        # Freeze header row
        ws_tx.freeze_panes = 'A2'

        logger.info(f"Added {len(year_transactions)} transactions to Verrichtingen sheet")

        # Create Aandachtspunten sheet if there are data quality warnings
        if report.has_data_quality_warnings:
            self._create_aandachtspunten_sheet(wb, report, header_fill, header_font, thin_border, currency_format)

        # Save workbook
        try:
            wb.save(output_path)
            logger.info(f"Successfully exported report to {output_path}")
        except Exception as e:
            logger.error(f"Failed to save Excel file to {output_path}: {e}")
            raise

    def _create_aandachtspunten_sheet(self, wb: Workbook, report: Report,
                                       header_fill, header_font, thin_border,
                                       currency_format: str) -> None:
        """Create Aandachtspunten sheet with data quality warnings.

        For MVP (US1): Shows uncategorized transaction details.
        """
        from openpyxl.styles import PatternFill

        ws = wb.create_sheet(title="Aandachtspunten")

        # Title
        ws.cell(row=1, column=1, value=f"Aandachtspunten - Boekjaar {report.fiscal_year}").font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        current_row = 3

        # Uncategorized transactions section (US1)
        if report.total_uncategorized != 0:
            count = len(report.uncategorized_items)
            total = float(report.total_uncategorized)

            # Warning summary
            warning_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            warning_cell = ws.cell(row=current_row, column=1,
                                   value=f"Niet-gecategoriseerde transacties: {count} (€ {total:,.2f})")
            warning_cell.font = Font(bold=True, color="92400E")
            warning_cell.fill = warning_fill
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 2

            # Filter uncategorized transactions
            uncategorized_tx = [
                t for t in self.transactions
                if t.booking_date.year == report.fiscal_year
                and not t.category
                and not t.is_excluded
            ]
            uncategorized_tx.sort(key=lambda t: t.booking_date)

            if uncategorized_tx:
                # Headers
                headers = ['Datum', 'Bedrag', 'Tegenpartij', 'Omschrijving']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                current_row += 1

                # Transaction rows
                for tx in uncategorized_tx:
                    ws.cell(row=current_row, column=1,
                            value=tx.booking_date.strftime('%d/%m/%Y')).border = thin_border
                    amount_cell = ws.cell(row=current_row, column=2, value=float(tx.amount))
                    amount_cell.number_format = currency_format
                    amount_cell.border = thin_border
                    ws.cell(row=current_row, column=3,
                            value=(tx.counterparty_name or '-')[:40]).border = thin_border
                    ws.cell(row=current_row, column=4,
                            value=(tx.description or '-')[:60]).border = thin_border
                    current_row += 1

                current_row += 1

            # Action required note
            ws.cell(row=current_row, column=1,
                    value="Actie: Controleer en categoriseer deze transacties.").font = Font(italic=True)
            current_row += 2

        # Verkeerde-rekening section (US2)
        if report.verkeerde_rekening_balance != 0:
            count = len(report.verkeerde_rekening_items)
            balance = float(report.verkeerde_rekening_balance)
            sign = "+" if balance > 0 else ""

            # Warning summary
            warning_fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
            warning_cell = ws.cell(row=current_row, column=1,
                                   value=f"Privé-uitgaven (verkeerde rekening) - Niet in balans: € {sign}{balance:,.2f}")
            warning_cell.font = Font(bold=True, color="5B21B6")
            warning_cell.fill = warning_fill
            ws.merge_cells(f'A{current_row}:D{current_row}')
            current_row += 2

            # Filter verkeerde-rekening transactions
            verkeerde_tx = [
                t for t in self.transactions
                if t.booking_date.year == report.fiscal_year
                and t.category == 'verkeerde-rekening'
                and not t.is_excluded
            ]
            verkeerde_tx.sort(key=lambda t: t.booking_date)

            if verkeerde_tx:
                # Headers
                headers = ['Datum', 'Bedrag', 'Tegenpartij', 'Omschrijving']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                current_row += 1

                # Transaction rows
                for tx in verkeerde_tx:
                    ws.cell(row=current_row, column=1,
                            value=tx.booking_date.strftime('%d/%m/%Y')).border = thin_border
                    amount_cell = ws.cell(row=current_row, column=2, value=float(tx.amount))
                    amount_cell.number_format = currency_format
                    amount_cell.border = thin_border
                    ws.cell(row=current_row, column=3,
                            value=(tx.counterparty_name or '-')[:40]).border = thin_border
                    ws.cell(row=current_row, column=4,
                            value=(tx.description or '-')[:60]).border = thin_border
                    current_row += 1

                current_row += 1

            # Action required note
            if balance < 0:
                action_text = "Actie: Voeg ontbrekende terugbetalingen toe of corrigeer de categorisatie."
            else:
                action_text = "Actie: Controleer of alle terugbetalingen correct zijn gecategoriseerd."
            ws.cell(row=current_row, column=1, value=action_text).font = Font(italic=True)
            current_row += 2

        # Adjust column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 50

        logger.info("Created Aandachtspunten sheet with data quality warnings")


    def format_for_console(self, report: Report) -> str:
        """
        Formats the P&L report for display in the console.
        """
        from src.services.depreciation import get_depreciation_for_year

        lines = []
        width = 55
        lines.append("=" * width)
        lines.append(f"   Resultatenrekening {report.fiscal_year}".center(width))
        lines.append("=" * width)
        lines.append("")

        # Income (Baten)
        lines.append("Baten")
        lines.append("-" * width)
        for item in report.income_items:
            lines.append(f"  {item.category:<40} € {item.amount:10,.2f}")
            for sub_item in item.sub_items:
                lines.append(f"    - {sub_item.category:<36} € {sub_item.amount:10,.2f}")
        lines.append("-" * width)
        lines.append(f"{'Totaal Baten':<42} € {report.total_income:10,.2f}")
        lines.append("")

        # Kosten (operational expenses - NOT depreciation)
        lines.append("Kosten")
        lines.append("-" * width)
        total_kosten = Decimal(0)
        for item in sorted(report.expense_items, key=lambda x: x.category):
            cat_name = item.category.replace('-', ' ').title()
            lines.append(f"  {cat_name:<40} € {-item.amount:10,.2f}")
            total_kosten += -item.amount
        lines.append("-" * width)
        lines.append(f"{'Totaal Kosten':<42} € {total_kosten:10,.2f}")
        lines.append("")

        # Afschrijvingen (Depreciation - separate from Kosten)
        if report.depreciation_items:
            lines.append("Afschrijvingen")
            lines.append("-" * width)
            depreciation_entries = get_depreciation_for_year(self.assets, report.fiscal_year)
            total_afschrijvingen = Decimal(0)
            for entry in depreciation_entries:
                asset_name = entry.asset_id
                for asset in self.assets:
                    if asset.id == entry.asset_id:
                        asset_name = asset.name
                        break
                lines.append(f"  {asset_name:<40} € {entry.amount:10,.2f}")
                total_afschrijvingen += entry.amount
            lines.append("-" * width)
            lines.append(f"{'Totaal Afschrijvingen':<42} € {total_afschrijvingen:10,.2f}")
            lines.append("")

        # Result
        lines.append("=" * width)
        profit_label = "Winst" if report.profit_loss >= 0 else "Verlies"
        lines.append(f"Resultaat ({profit_label}){'':<22} € {report.profit_loss:10,.2f}")
        lines.append("=" * width)
        lines.append("")

        # Verworpen uitgaven
        if report.disallowed_expenses:
            lines.append("Verworpen Uitgaven (Disallowed Expenses)")
            lines.append("-" * width)
            for item in sorted(report.disallowed_expenses, key=lambda x: x.category):
                lines.append(f"  {item.category:<25} € {item.total_amount:8,.2f} ({item.deductible_pct}% aftrekbaar)")
                lines.append(f"    Verworpen: € {item.disallowed_amount:8,.2f}")
            lines.append("-" * width)
            lines.append(f"{'Totaal Verworpen':<37} € {report.total_disallowed:10,.2f}")
            lines.append("")

        # Uncategorized warning
        if report.total_uncategorized != 0:
            lines.append(f"Waarschuwing: Er is € {report.total_uncategorized:,.2f} aan niet-gecategoriseerde transacties.")
            lines.append("Het resultaat is exclusief deze transacties.")

        return "\n".join(lines)
